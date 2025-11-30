import os
from openpyxl import Workbook


def save_wb(wb: Workbook, path: str):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    print(f"Wrote {path}")


def make_batches_xlsx(path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "batches"
    ws.append(["id", "from", "to"])  # headers
    for year in range(2022, 2033):
        ws.append([f"B{year}", f"{year}", f"{year+1}"])
    save_wb(wb, path)


def make_doctors_xlsx(path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "doctors"
    ws.append(["id", "name", "department", "title", "phone", "email"])  # headers
    departments = ["Surgery", "Orthodontics", "Pediatrics", "Endodontics", "Prosthodontics"]
    titles = ["Doctor", "Teaching Assistant"]
    for i in range(1, 101):
        dep = departments[i % len(departments)]
        title = titles[i % len(titles)]
        ws.append([f"FM{i:03d}", f"Dr. Name {i}", dep, title, f"0100000{i:04d}", f"doctor{i}@example.com"]) 
    save_wb(wb, path)


def make_students_xlsx(path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "students"
    ws.append(["id", "name", "batch", "department", "round", "phone", "email"])  # headers
    departments = ["Surgery", "Orthodontics", "Pediatrics", "Endodontics", "Prosthodontics"]
    batches = [f"B{y}" for y in range(2022, 2033)]
    rounds = [f"R{i:03d}" for i in range(1, 25)]
    for i in range(1, 501):
        ws.append([
            f"stu{i:03d}",
            f"Student {i}",
            batches[i % len(batches)],
            departments[i % len(departments)],
            rounds[i % len(rounds)],
            f"0101{i:07d}",
            f"student{i}@example.com"
        ])
    save_wb(wb, path)


def make_departments_xlsx(path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "departments"
    ws.append(["id", "name", "manager", "capacity", "description"])  # manager = faculty_members_id
    names = [
        ("d001", "Surgery", "FM001", 60, "General surgery department"),
        ("d002", "Orthodontics", "FM002", 45, "Braces and alignment"),
        ("d003", "Pediatrics", "FM003", 35, "Children dentistry"),
        ("d004", "Endodontics", "FM004", 30, "Root canal treatments"),
        ("d005", "Prosthodontics", "FM005", 40, "Dental prostheses")
    ]
    for row in names:
        ws.append(list(row))
    save_wb(wb, path)


def make_rounds_xlsx(path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "rounds"
    ws.append(["name", "batch", "month"])  # rounds importer uses these keys
    months = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
    for year in range(2022, 2033):
        for m in months:  # full year
            ws.append([f"{m} {year}", f"B{year}", m])
    save_wb(wb, path)


def main():
    base = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "mock_data"))
    make_batches_xlsx(os.path.join(base, "batches.xlsx"))
    make_doctors_xlsx(os.path.join(base, "doctors.xlsx"))
    make_students_xlsx(os.path.join(base, "students.xlsx"))
    make_departments_xlsx(os.path.join(base, "departments.xlsx"))
    make_rounds_xlsx(os.path.join(base, "rounds.xlsx"))


if __name__ == "__main__":
    main()


